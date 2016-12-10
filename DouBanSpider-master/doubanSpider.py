#-*- coding: UTF-8 -*-

import importlib,sys 
import time
import re
import urllib.request
import urllib.error 
import numpy as np  
from bs4 import BeautifulSoup
from openpyxl import Workbook

importlib.reload(sys)


hds=['Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6',\
'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11',\
'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)']
length = len(hds)

def book_spider(book_tag,page_num):
    num=0;
    book_list=[]
    try_times=0
    
    while num < int(page_num):
        url='http://www.douban.com/tag/'+urllib.parse.quote(book_tag)+'/book?start='+str(num*15)
        time.sleep(np.random.rand()*5)
        
        #Last Version
        try:
            req = urllib.request.Request(url)
            req.add_header('User-Agent',hds[num%length])
            response = urllib.request.urlopen(req)
            plain_text=response.read() 
        except (urllib.error.HTTPError, urllib.error.URLError) as e:
            print('对不起，没有%s这个标签' %book_tag)
            continue
        
        soup = BeautifulSoup(plain_text,"lxml") 
        list_soup = soup.find('div', {'class': 'mod book-list'})
        
        try_times+=1;
        if list_soup==None and try_times<200:
            continue
        elif list_soup==None or len(list_soup)<=1:
            break # Break when no informatoin got after 200 times requesting
        
        for book_info in list_soup.findAll('dd'):
            title = book_info.find('a', {'class':'title'}).string.strip()
            desc = book_info.find('div', {'class':'desc'}).string.strip()
            desc_list = desc.split('/')
            book_url = book_info.find('a', {'class':'title'}).get('href')
            
            try:
                author_info = '作者/译者： ' + '/'.join(desc_list[0:-3])
            except:
                author_info ='作者/译者： 暂无'
            try:
                pub_info = '出版信息： ' + '/'.join(desc_list[-3:])
            except:
                pub_info = '出版信息： 暂无'
            try:
                rating = book_info.find('span', {'class':'rating_nums'}).string.strip()
            except:
                rating='0.0'
            try:
                people_num = get_people_num(book_url)
                people_num = people_num.strip('人评价')
            except:
                people_num ='0'
            
            book_list.append([title,rating,people_num,author_info,pub_info])
            try_times=0 #set 0 when got valid information
            
        num+=1
        print('Downloading Information From Page %d' % num)
        
       
    return book_list


def get_people_num(url):
    try:
        req = urllib.request.Request(url)
        req.add_header('User-Agent',hds[np.random.randint(0,len(hds))])
        response = urllib.request.urlopen(req)
        plain_text=response.read() 
    except (urllib.error.HTTPError, urllib.error.URLError) as e:
        print(e)
    soup = BeautifulSoup(plain_text,"lxml")
    people_num=soup.find('div',{'class':'rating_sum'}).findAll('span')[1].string.strip()
    return people_num


def do_spider(book_tag_lists,page_num=1):
    book_lists=[]
    for book_tag in book_tag_lists:
        book_list=book_spider(book_tag,page_num)
        book_list=sorted(book_list,key=lambda x:x[1],reverse=True) 
        book_lists.append(book_list)
    return book_lists


def print_book_lists_excel(book_lists,book_tag_lists):
    #wb=Workbook(optimized_write=True)
    wb=Workbook()
    ws=[]
    for i in range(len(book_tag_lists)):
        ws.append(wb.create_sheet(title=book_tag_lists[i])) 
    for i in range(len(book_tag_lists)): 
        ws[i].append(['序号','书名','评分','评价人数','作者','出版社'])
        count=1
        for bl in book_lists[i]:
            ws[i].append([count,bl[0],float(bl[1]),int(bl[2]),bl[3],bl[4]])
            count+=1
    save_path='book_list'
    for i in range(len(book_tag_lists)):
        save_path+=('-'+book_tag_lists[i])
    save_path+='.xlsx'
    wb.save(save_path)

    
if __name__=='__main__':
    module = input("请选择您需要爬取的具体几个模块(如计算机，心理)：")
    page_num = input("请输入您要查找的页数（从当前页开始）：")
    book_tag_lists = module.split('，')
    book_lists=do_spider(book_tag_lists,page_num)
    print_book_lists_excel(book_lists,book_tag_lists)
    
