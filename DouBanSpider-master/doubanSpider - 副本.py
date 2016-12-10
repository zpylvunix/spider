#-*- coding: UTF-8 -*-

import importlib,sys 
import time
import re
import urllib.request
import urllib.error 
import numpy as np  
from bs4 import BeautifulSoup
from openpyxl import Workbook

importlib.reload(sys)


hds=['Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6',\
'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11',\
'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)']
length = len(hds)

def book_spider(book_tag,page_num):
    num=0;
    book_list=[]
    try_times=0
    
    while num < int(page_num):
	#按照标准， URL 只允许一部分 ASCII 字符（数字字母和部分符号），其他的字符（如汉字）是不符合 URL 标准的。
	#所以对于不符合url标准的那部分字符需要用urllib.parse.quote进行编码。
        '''
	URL编码的方式是把需要编码的字符转化为 %xx 的形式。通常 URL 编码是基于 UTF-8 的（当然这和浏览器平台有关）。
        例子：比如『我』，unicode 为 0x6211,  UTF-8 编码为 0xE6 0x88 0x91，URL 编码就是 %E6%88%91
	还有作用就是屏蔽空格
	'''
        #https://www.douban.com/tag/%E5%B0%8F%E8%AF%B4/book?start=0
        url='http://www.douban.com/tag/'+urllib.parse.quote(book_tag)+'/book?start='+str(num)
        time.sleep(np.random.rand()*5)  #产生符合正态分布的随机数（0-1区间）
        
        #Last Version
        try:
            req = urllib.request.Request(url)
            req.add_header('User-Agent',hds[num%length])
            response = urllib.request.urlopen(req)
            plain_text=response.read() 
        except (urllib.error.HTTPError, urllib.error.URLError) as e:
            print(e)
            continue
  
        
        soup = BeautifulSoup(plain_text,"lxml")  #创建一个BeautifulSoup对象
        list_soup = soup.find('div', {'class': 'mod book-list'})
        
        try_times+=1;
        if list_soup==None and try_times<200:
            continue
        elif list_soup==None or len(list_soup)<=1:
            break # Break when no informatoin got after 200 times requesting
        
        for book_info in list_soup.findAll('dd'): #书的列表 book_info：<class 'bs4.element.Tag'>
            title = book_info.find('a', {'class':'title'}).string.strip()
            desc = book_info.find('div', {'class':'desc'}).string.strip()
            desc_list = desc.split('/')
            book_url = book_info.find('a', {'class':'title'}).get('href') #根据href键得到值，为了获得评价人数的连接
            
            try:
                author_info = '作者/译者： ' + '/'.join(desc_list[0:-3])
            except:
                author_info ='作者/译者： 暂无'
            try:
                pub_info = '出版信息： ' + '/'.join(desc_list[-3:])
            except:
                pub_info = '出版信息： 暂无'
            try:
                rating = book_info.find('span', {'class':'rating_nums'}).string.strip()  #评分
            except:
                rating='0.0'
            try:
                people_num = get_people_num(book_url)
                people_num = people_num.strip('人评价')
            except:
                people_num ='0'

            book_list.append([title,rating,people_num,author_info,pub_info])   #在列表追加列表元素 两重列表
            try_times=0 #set 0 when got valid information
        num+=1
        print('Downloading Information From Page %d' % num)
    return book_list

#得到评价人数 get
def get_people_num(url):
    try:
        req = urllib.request.Request(url)
        req.add_header('User-Agent',hds[np.random.randint(0,len(hds))])
        response = urllib.request.urlopen(req)
        plain_text=response.read() 
    except (urllib.error.HTTPError, urllib.error.URLError) as e:
        print(e)
    soup = BeautifulSoup(plain_text,"lxml")
    people_num=soup.find('div',{'class':'rating_sum'}).findAll('span')[1].string.strip()
    return people_num


def do_spider(book_tag_lists,page_num=1):
    book_lists=[]
    for book_tag in book_tag_lists:
        book_list=book_spider(book_tag,page_num)  #循环，指定要爬取的当前模块和页数
        book_list=sorted(book_list,key=lambda x:x[1],reverse=True)  # sorted(iterable, cmp=None, key=None, reverse=False) False:代表升序 根据分数排名
        book_lists.append(book_list)  #三重列表  第一重书的信息  第二重单个模块的所有书的信息 第三重所有模块书的信息
    return book_lists


def print_book_lists_excel(book_lists,book_tag_lists):
    #wb=Workbook(optimized_write=True)
    wb=Workbook()
    ws=[]
    for i in range(len(book_tag_lists)):
        ws.append(wb.create_sheet(title=book_tag_lists[i]))   #根据多个模块名模块名创建表格   表格和列表就联系上了
    for i in range(len(book_tag_lists)): 
        ws[i].append(['序号','书名','评分','评价人数','作者','出版社'])   #ws[i]表格i
        count=1  #序号
        for bl in book_lists[i]:  #单个模块book_lists[i]，双重列表  b1就是单重列表，表示书的信息
            ws[i].append([count,bl[0],float(bl[1]),int(bl[2]),bl[3],bl[4]])
            count+=1
    #构建函数名  get
    save_path='book_list'
    for i in range(len(book_tag_lists)):
        save_path+=('-'+book_tag_lists[i])
    save_path+='.xlsx'
    wb.save(save_path)

#类列表
def get_kindlist():
    url = 'https://www.douban.com/tag/'
    try:
        req = urllib.request.Request(url)
        req.add_header('User-Agent',hds[np.random.randint(0,len(hds))])
        response = urllib.request.urlopen(req)
        plain_text=response.read() 
    except (urllib.error.HTTPError, urllib.error.URLError) as e:
        print(e)
    soup = BeautifulSoup(plain_text,"lxml")
    list=soup.find('div',{'class':'article'}).findAll('h2')
    kindlist = []
    for element in list :
        kindlist.append(element.string)
    print(kindlist)
    
    

#模块列表
def get_modulelist(kind):
    url = 'https://www.douban.com/tag/'
    try:
        req = urllib.request.Request(url)
        req.add_header('User-Agent',hds[np.random.randint(0,len(hds))])
        response = urllib.request.urlopen(req)
        plain_text=response.read() 
    except (urllib.error.HTTPError, urllib.error.URLError) as e:
        print(e)
    soup = BeautifulSoup(plain_text,"lxml")
    list=soup.find('div',{'class':'article'}).findAll('h2')
    i=0
    for element in list :
        if kind != element.string:
           i = i+1
        else :
            break
    list1 =soup.find('div',{'class':'article'}).findAll('ul',{'class':'topic-list'})[i]
    soup1 = BeautifulSoup(str(list1),"lxml")
    list2 =soup1.findAll('a')
    modulelist = []
    for element in list2:
        modulelist.append(element.string)
    
    print(modulelist)
    
if __name__=='__main__':
    get_kindlist()
    kind = input("请从上面分类中选择你感兴趣的一种(如美食)：")
    get_modulelist(kind)
    module = input("请选择该类下的中具体几个模块(如便当，蛋糕)：")
    page_num = input("请输入您要查找的页数（从当前页开始）：")
    book_tag_lists = module.split('，')
    book_lists=do_spider(book_tag_lists,page_num)
    print_book_lists_excel(book_lists,book_tag_lists)
    
